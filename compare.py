"""
PE Firm Weekly Team Comparison — compare.py
============================================
Compares two weekly employee snapshots and produces a change report.

Usage:
    python compare.py                          # auto-detects the two latest employees_*.xlsx files
    python compare.py PREV.xlsx CURR.xlsx      # explicit file paths
    python compare.py --prev PREV.xlsx --curr CURR.xlsx

Output:
    report_YYYYMMDD_HHMMSS.xlsx with two sheets:
      "Current Employees"  — all current rows + "change" column, yellow-highlighted positions that changed
      "Leavers"            — employees from previous week who are no longer present
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import argparse
import re
from datetime import datetime
from glob import glob
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ────────────────────────────────────────────────────────────────

YELLOW     = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
LIGHT_RED  = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
HEADER_BG  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FG  = Font(bold=True, color="FFFFFF", size=11)
BOLD       = Font(bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def normalise(text) -> str:
    """Normalise a string for comparison: lower, collapse whitespace, strip."""
    if not text or (isinstance(text, float)):
        return ""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def match_key(firm: str, name: str) -> str:
    """Composite match key: firm + person name, both normalised."""
    return f"{normalise(firm)}||{normalise(name)}"


def auto_detect_files() -> tuple[str, str]:
    """Find the two most-recent employees_*.xlsx files in the working directory."""
    files = sorted(glob("employees_*.xlsx"))
    if len(files) < 2:
        print(f"ERROR: Need at least 2 employees_*.xlsx files, found {len(files)}: {files}")
        sys.exit(1)
    prev, curr = files[-2], files[-1]
    print(f"Auto-detected files:\n  PREV: {prev}\n  CURR: {curr}")
    return prev, curr


def load(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    # Normalise column names in case of minor casing differences
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    # Ensure required columns exist
    for col in ["firm_name", "person_name", "person_position"]:
        if col not in df.columns:
            raise ValueError(f"Missing required column '{col}' in {path}")
    # Fill NaN with empty string for comparison, keep N/A as-is
    df["firm_name"]       = df["firm_name"].fillna("").astype(str).str.strip()
    df["person_name"]     = df["person_name"].fillna("").astype(str).str.strip()
    df["person_position"] = df["person_position"].fillna("N/A").astype(str).str.strip()
    return df


# ── Core comparison ──────────────────────────────────────────────────────────

def compare(prev_df: pd.DataFrame, curr_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Compare previous and current datasets.

    Returns
    -------
    curr_with_changes : DataFrame
        All rows from curr_df with an extra 'change' column:
          ""          — no change
          "New Hire"  — person not in prev
          "Promotion" — person present but position changed
    leavers : DataFrame
        Rows from prev_df for employees no longer in curr_df.
    """
    # Build lookup: key → row (dict) for each dataset
    prev_lookup: dict[str, dict] = {}
    for _, row in prev_df.iterrows():
        k = match_key(row["firm_name"], row["person_name"])
        if k and k not in prev_lookup:        # keep first occurrence per key
            prev_lookup[k] = row.to_dict()

    curr_lookup: dict[str, dict] = {}
    for _, row in curr_df.iterrows():
        k = match_key(row["firm_name"], row["person_name"])
        if k and k not in curr_lookup:
            curr_lookup[k] = row.to_dict()

    # ── Tag current rows ────────────────────────────────────────────────────
    changes        = []   # ("", "New Hire", "Promotion") per curr row
    prev_positions = []   # previous position for Promotion rows, else ""

    for _, row in curr_df.iterrows():
        k = match_key(row["firm_name"], row["person_name"])
        if not k:
            changes.append("")
            prev_positions.append("")
            continue

        if k not in prev_lookup:
            changes.append("New Hire")
            prev_positions.append("")
        else:
            prev_pos = normalise(prev_lookup[k].get("person_position", ""))
            curr_pos = normalise(row["person_position"])
            if prev_pos and curr_pos and prev_pos != curr_pos:
                changes.append("Promotion")
                prev_positions.append(prev_lookup[k].get("person_position", ""))
            else:
                changes.append("")
                prev_positions.append("")

    curr_with_changes = curr_df.copy()
    curr_with_changes["change"]        = changes
    curr_with_changes["previous_role"] = prev_positions

    # ── Find leavers ────────────────────────────────────────────────────────
    leaver_rows = []
    for k, row in prev_lookup.items():
        if k not in curr_lookup:
            leaver_rows.append({
                "firm_name":           row.get("firm_name", ""),
                "person_name":         row.get("person_name", ""),
                "last_known_position": row.get("person_position", "N/A"),
                "last_seen_date":      row.get("date_scraped", ""),
            })

    leavers = pd.DataFrame(leaver_rows, columns=[
        "firm_name", "person_name", "last_known_position", "last_seen_date"
    ])

    return curr_with_changes, leavers


# ── Excel formatting ─────────────────────────────────────────────────────────

def _style_header_row(ws, row_num: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_BG
        cell.font   = HEADER_FG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autofit_columns(ws, min_w=10, max_w=50):
    for col_cells in ws.iter_cols():
        max_len = 0
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


def write_report(curr_changes: pd.DataFrame, leavers: pd.DataFrame, output_path: str):
    """Write the two-sheet Excel report with highlights."""

    # ── Decide which columns to write ───────────────────────────────────────
    display_cols = [
        "firm_name", "person_name", "person_position", "previous_role",
        "change", "team", "location", "date_scraped",
    ]
    # Only include columns that actually exist
    display_cols = [c for c in display_cols if c in curr_changes.columns]

    leaver_cols = ["firm_name", "person_name", "last_known_position", "last_seen_date"]

    # Write initial data with pandas (handles dtypes cleanly)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        curr_changes[display_cols].to_excel(
            writer, sheet_name="Current Employees", index=False
        )
        leavers[leaver_cols].to_excel(
            writer, sheet_name="Leavers", index=False
        )

    # Re-open with openpyxl for formatting
    wb = load_workbook(output_path)

    # ── Sheet 1: Current Employees ──────────────────────────────────────────
    ws1 = wb["Current Employees"]
    _style_header_row(ws1, 1, len(display_cols))

    # Locate column indices (1-based)
    header_map = {cell.value: cell.column for cell in ws1[1]}
    pos_col    = header_map.get("person_position")
    change_col = header_map.get("change")
    prev_col   = header_map.get("previous_role")

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        change_val = ""
        if change_col:
            change_val = str(row[change_col - 1].value or "")

        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)

        if change_val == "Promotion":
            # Yellow on the position cell that changed
            if pos_col:
                row[pos_col - 1].fill = YELLOW
                row[pos_col - 1].font = BOLD
        elif change_val == "New Hire":
            # Light blue on the entire row
            for cell in row:
                cell.fill = LIGHT_BLUE

    # Freeze the header row
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Leavers ────────────────────────────────────────────────────
    ws2 = wb["Leavers"]
    _style_header_row(ws2, 1, len(leaver_cols))

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")
        # Light red background for all leaver rows
        for cell in row:
            cell.fill = LIGHT_RED

    ws2.freeze_panes = "A2"

    _autofit_columns(ws1)
    _autofit_columns(ws2)

    wb.save(output_path)
    print(f"Report saved → {output_path}")


# ── Summary print ────────────────────────────────────────────────────────────

def print_summary(curr_changes: pd.DataFrame, leavers: pd.DataFrame,
                  prev_path: str, curr_path: str):
    total_curr  = len(curr_changes)
    promotions  = (curr_changes["change"] == "Promotion").sum()
    new_hires   = (curr_changes["change"] == "New Hire").sum()
    unchanged   = total_curr - promotions - new_hires
    n_leavers   = len(leavers)

    print()
    print("=" * 55)
    print("  WEEKLY CHANGE SUMMARY")
    print("=" * 55)
    print(f"  Previous snapshot : {Path(prev_path).name}")
    print(f"  Current snapshot  : {Path(curr_path).name}")
    print("-" * 55)
    print(f"  Current employees : {total_curr:>6,}")
    print(f"    Unchanged        : {unchanged:>6,}")
    print(f"    Role changes     : {promotions:>6,}  (highlighted yellow)")
    print(f"    New hires        : {new_hires:>6,}  (highlighted blue)")
    print(f"  Leavers           : {n_leavers:>6,}  (Leavers tab, red)")
    print("=" * 55)

    if promotions > 0:
        print(f"\n  Top role changes (first 10):")
        promo_rows = curr_changes[curr_changes["change"] == "Promotion"]
        for _, r in promo_rows.head(10).iterrows():
            print(f"    [{r['firm_name']}] {r['person_name']}")
            print(f"       {r.get('previous_role', '?')}  →  {r['person_position']}")

    if n_leavers > 0:
        print(f"\n  Leavers by firm:")
        for firm, cnt in leavers["firm_name"].value_counts().head(10).items():
            print(f"    {firm}: {cnt}")

    print()


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Compare two weekly PE employee snapshots and produce a change report."
    )
    parser.add_argument("files", nargs="*",
                        help="Optional: PREV.xlsx CURR.xlsx as positional args")
    parser.add_argument("--prev", help="Previous week's Excel file")
    parser.add_argument("--curr", help="Current week's Excel file")
    parser.add_argument("--output", help="Output report filename (default: auto-named)")
    args = parser.parse_args()

    # Resolve input files
    if args.prev and args.curr:
        prev_path, curr_path = args.prev, args.curr
    elif len(args.files) == 2:
        prev_path, curr_path = args.files
    elif len(args.files) == 0:
        prev_path, curr_path = auto_detect_files()
    else:
        parser.print_help()
        sys.exit(1)

    # Output filename
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or f"report_{timestamp}.xlsx"

    print(f"\nLoading PREV: {prev_path}")
    prev_df = load(prev_path)
    print(f"  → {len(prev_df):,} rows, {prev_df['firm_name'].nunique()} firms")

    print(f"Loading CURR: {curr_path}")
    curr_df = load(curr_path)
    print(f"  → {len(curr_df):,} rows, {curr_df['firm_name'].nunique()} firms")

    print("\nComparing...")
    curr_changes, leavers = compare(prev_df, curr_df)

    print_summary(curr_changes, leavers, prev_path, curr_path)
    write_report(curr_changes, leavers, output_path)


if __name__ == "__main__":
    main()
